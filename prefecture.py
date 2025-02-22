from selenium import webdriver
from selenium.webdriver.chrome import service
from selenium.webdriver.common.by import By
import time  # 不使用
from selenium.webdriver.common.keys import Keys
import bs4
import openpyxl

url="https://www.google.co.jp"
# シート読み込み
excel_file='都道府県.xlsx'
book = openpyxl.load_workbook(excel_file)
sheet = book['シート1']

# ChromeDriverのパスを変数に設定
CHROMEDRIVER = "\chromedriver\chromedriver.exe"
# ChromeDriverのstartとstopを制御するServiceオブジェクトを介してパスを渡す
chrome_service = service.Service(executable_path=CHROMEDRIVER)

#途中で止まる
for i in range(1,2):
    # Chromeを起動
    driver = webdriver.Chrome(service=chrome_service)
    # 指定したURLに遷移する
    driver.get(url)
    # 検索テキストボックスの要素をname属性名から取得
    element = driver.find_element(By.NAME, "q")
    # 検索テキストボックスに文字列を入力
    element.send_keys(sheet.cell(row=i, column=1).value)
    element.send_keys(Keys.ENTER)

    html = driver.page_source
    soup = bs4.BeautifulSoup(html, "html.parser")
    word = soup.find('h2', class_='qrShPb kno-ecr-pt PZPZlf q8U8x')
    sheet.cell(row=i,column=2,value=word.string)
    book.save('都道府県.xlsx')
    driver.close()
